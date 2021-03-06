#!/usr/bin/python
#-*- coding: utf-8 -*-
import os
import xlrd
import openpyxl
import datetime

input_Folder = "/home/quocdai/Projects/input"
output_Folder = "/home/quocdai/Projects/output"
done_Folder = "/home/quocdai/Projects/done"
teacher_Name = {'MTCNHUNG':'Mai Thị Cẩm Nhung','VHKHANH':'Võ Hồng Khanh','TCNGHI':'Trần Công Nghị','NVTTHAI':'Nguyễn Võ Thông Thái',\
'MQDAI':'Mai Quốc Đại','VDANH':'Võ Duy Anh','LHHUONG':'Lương Hoàng Hướng'}

#input_file = "/home/quocdai/Downloads/Overview.xls"
teacher_original_file = "/home/quocdai/Downloads/GV.xlsx"
service_original_file = "/home/quocdai/Downloads/DV.xlsx"


def get_inputFile():
    return os.listdir(input_Folder)

def class_Info(inputFile):
    return inputFile.split('_')
    
def count_GP(GP_column=0):
    workbook_to_read = xlrd.open_workbook(inputFile,on_demand=True)
    
    #First sheet
    sheet = workbook_to_read.sheet_by_index(0) 
    
    #Get list of value in column
    col_value = sheet.col_values(GP_column, 1, sheet.nrows - 2) 
    
    #Just get firest number in celvalue, Damn Thai! 
    col_value = [str(x)[0] for x in col_value]
    
    
    return (col_value.count('4'),col_value.count('3'),col_value.count('2'),col_value.count('1'))

def read_Remark():
    workbook_to_read = xlrd.open_workbook(inputFile,on_demand=True)
    
    #First sheet
    sheet = workbook_to_read.sheet_by_index(0) 
    
    
    #Get list of value in column
    list_Remark = sheet.col_values(26, 1, sheet.nrows - 2) 
    list_Remark = [x for x in list_Remark if x != '--']
    list_Remark = [x for x in list_Remark if x != 'Requires grading' ]
    return list_Remark
       

def write_Teacher_GP(output_File):
    write_workbook = openpyxl.load_workbook(teacher_original_file)    
    write_workbook.sheetnames
    sheet1 = write_workbook['sheet1']
    
    #Read and insert data to GP.
    for record in range (0,11):
        list_GP = count_GP(record+5) #GP teacher start from comlumn 5 to 15 in file
        for GP in range (0,4):
            sheet1[chr(ord('F')+GP)+str(17+record)] = list_GP[GP]
    
    
     #Write remark
    if len(read_Remark()) != 0:
        remark ="'"
        for line in read_Remark():
            remark = remark+"-"+line+"\n"
        sheet1['B31'] = remark[0:-1]

        
    #Write basic class info
    sheet1['B10'] = teacher_Name[class_Info(inputFile)[4]]
    sheet1['B11'] = class_Info(inputFile)[6][6:8]+"/"+class_Info(inputFile)[6][4:6]+"/"+class_Info(inputFile)[6][0:4]
    sheet1['G11'] = class_Info(inputFile)[1]
    sheet1['G10'] = class_Info(inputFile)[2]
    sheet1['E11'] = class_Info(inputFile)[3]
    sheet1['K10'] = class_Info(inputFile)[5]
    sheet1['J33'] = datetime.datetime.now().strftime("%d/%m/%Y")
    sheet1['J40'] = datetime.datetime.now().strftime("%d/%m/%Y")

    #Write logo image and data to output file    
    img = openpyxl.drawing.image.Image("cusc.png")
    sheet1.add_image(img, 'A1')
    write_workbook.save(output_File)

def write_Service_GP(output_File):
    write_workbook = openpyxl.load_workbook(service_original_file)    
    write_workbook.sheetnames
    sheet1 = write_workbook['sheet1']

    #Read and insert data to GP.
    for record in range (0,10):
        list_GP = count_GP(record+16) #GP service Start from comlumn 1 to 26 in file
        for GP in range (0,4):
            sheet1[chr(ord('F')+GP)+str(17+record)] = list_GP[GP]
    
       
    #Write basic class info
    sheet1['B10'] = class_Info(inputFile)[6][6:8]+"/"+class_Info(inputFile)[6][4:6]+"/"+class_Info(inputFile)[6][0:4]
    sheet1['G11'] = class_Info(inputFile)[1]
    sheet1['G10'] = class_Info(inputFile)[3]
    sheet1['K10'] = class_Info(inputFile)[5]
    sheet1['J32'] = datetime.datetime.now().strftime("%d/%m/%Y")
    sheet1['J38'] = datetime.datetime.now().strftime("%d/%m/%Y")
    
    
    #Write logo image and data to output file
    img = openpyxl.drawing.image.Image("cusc.png")
    sheet1.add_image(img, 'A1')
    write_workbook.save(output_File)

def write_Remark(output_File):
    if len(read_Remark()) != 0:
        remark_File = open(output_File,"w+")
        for line in read_Remark():
            remark_File.write("-"+line.encode('utf8')+"\n")
        remark_File.close()

#write_Teacher_GP()
#write_Service_GP()

def run():
    global inputFile
    for iFile in get_inputFile():
        inputFile = input_Folder + "/" + iFile
        doneFile = done_Folder + "/" + iFile + ".done"
        DV_output_File = output_Folder + "/" + iFile[0:-12] + "DV_" + iFile[-12:] + "x"
        GV_output_File = output_Folder + "/" + iFile[0:-12] + "GV_" + iFile[-12:] + "x"
        RM_output_File = output_Folder + "/" + iFile[0:-12] + "DV_" + iFile[-12:-3] + "txt"
        
        write_Service_GP(DV_output_File)
        write_Teacher_GP(GV_output_File)
        write_Remark(RM_output_File)
        os.rename(inputFile,doneFile)
             
run()



#    print "Date: "+class_Info(xfile)[6][6:8]+"/"+class_Info(xfile)[6][4:6]+"/"+class_Info(xfile)[6][0:4]
#    print "Batch: "+class_Info(xfile)[1]
#    print "Subject: "+class_Info(xfile)[2]
#    print "Course: "+class_Info(xfile)[3]
#    print "Faculty: "+class_Info(xfile)[4]
#    print "Room: "+class_Info(xfile)[5]
#    print "-------------------"
#    print datetime.datetime.now().strftime("%d/%m/%Y")
